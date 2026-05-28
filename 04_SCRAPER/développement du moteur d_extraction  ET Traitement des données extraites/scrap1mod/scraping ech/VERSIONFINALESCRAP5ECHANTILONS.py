"""
============================================================
PROJET G1/G2 — Scraper Kompass V8 (Version Définitive F-2049)
============================================================
- Bypass Cloudflare (DrissionPage)
- Ciblage DOM précis (Anti-bug des "2 millions")
- Base SQL normalisée T1.1.4
- Export Excel natif typé pour Power BI (.xlsx)
============================================================
"""

import sqlite3
import random
import logging
import re
import time
import os
from datetime import datetime
from urllib.parse import urlparse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from DrissionPage import ChromiumPage, ChromiumOptions

# ============================================================
# CONFIGURATION
# ============================================================
CONFIG = {
    "excel_input":   "liste_URL_KOMPASS.xlsx",
    "db_path":       "base_reindustrialisation_test.db",
    "excel_output":  "Data_PowerBI_Propre.xlsx",
    "log_path":      "scraper_v8.log",
    "delay_min":     4.0,
    "delay_max":     8.0,
    "max_retries":   2,
    "taille_test":   5, # Test sur 5 entreprises
}

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", handlers=[logging.StreamHandler()])
log = logging.getLogger("KompassV8")

def trouver_navigateur():
    chemins = [
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"
    ]
    for c in chemins:
        if os.path.exists(c): return c
    return None

# ============================================================
# BASE DE DONNÉES (Structure exacte F-2049)
# ============================================================
def init_db(path):
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS entreprises (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            lien_kompass        TEXT UNIQUE,
            siret               TEXT,
            siren               TEXT,
            code_naf            TEXT,
            raison_sociale      TEXT NOT NULL,
            forme_juridique     TEXT,
            date_creation       TEXT,
            adresse             TEXT,
            ville               TEXT,
            code_postal         TEXT,
            departement         TEXT,
            region              TEXT,
            site_web            TEXT,
            telephone           TEXT,
            email               TEXT,
            chiffre_affaires    REAL,
            effectifs           INTEGER,
            description         TEXT,
            secteur_ia          TEXT,
            filiere_ia          TEXT,
            date_scraping       TEXT,
            statut_scraping     TEXT DEFAULT 'pending',
            html_brut           TEXT
        );
    """)
    conn.commit()
    return conn

def sauvegarder_entreprise(conn, data):
    conn.execute("""
        INSERT INTO entreprises (
            lien_kompass, raison_sociale, siret, siren, code_naf, forme_juridique, 
            date_creation, adresse, ville, code_postal, site_web, telephone, email, 
            chiffre_affaires, effectifs, description, date_scraping, statut_scraping, html_brut
        ) VALUES (
            :lien_kompass, :raison_sociale, :siret, :siren, :code_naf, :forme_juridique, 
            :date_creation, :adresse, :ville, :code_postal, :site_web, :telephone, :email, 
            :chiffre_affaires, :effectifs, :description, :date_scraping, :statut_scraping, :html_brut
        )
        ON CONFLICT(lien_kompass) DO UPDATE SET
            telephone=excluded.telephone, email=excluded.email, adresse=excluded.adresse,
            chiffre_affaires=excluded.chiffre_affaires, effectifs=excluded.effectifs,
            siren=excluded.siren, siret=excluded.siret, code_naf=excluded.code_naf,
            statut_scraping=excluded.statut_scraping, date_scraping=excluded.date_scraping
    """, data)
    conn.commit()

# ============================================================
# NETTOYAGE NUMÉRIQUE POUR POWER BI
# ============================================================
def nettoyer_chiffre_affaires(ca_str):
    if not ca_str: return None
    ca_lower = ca_str.lower()
    num_str = re.sub(r'[^\d,.]', '', ca_str).replace(',', '.')
    
    try:
        val = float(num_str)
        if re.search(r'\bm€?\b|\bmeur\b|\bmillions?\b|\bm\b|\bm €\b', ca_lower):
            val *= 1_000_000
        elif re.search(r'\bk€?\b|\bkeur\b|\bmilliers?\b|\bk\b|\bk €\b', ca_lower):
            val *= 1_000
        return val
    except:
        return None

def nettoyer_effectifs(eff_str):
    if not eff_str: return None
    num = re.sub(r'[^\d]', '', eff_str)
    try:
        return int(num)
    except:
        return None

# ============================================================
# EXTRACTION (VISÉE LASER)
# ============================================================
def extraire(page, meta):
    html = page.html

    def clean_text(text):
        if not text: return ""
        return re.sub(r'\s+', ' ', text).strip()

    def find_regex(pattern, defaut=""):
        try:
            m = re.search(pattern, html, re.I | re.S)
            if m: return clean_text(m.group(1) if m.groups() else m.group(0))
        except: pass
        return defaut

    siret_brut = find_regex(r"SIRET[^:]*:\s*([\d\s]{14,20})")
    siret = re.sub(r"\s", "", siret_brut)[:14] if siret_brut else ""
    siren = siret[:9] if siret else re.sub(r"\D", "", find_regex(r"\b(\d{9})\b"))[:9]

    telephone = ""
    try:
        tel_link = page.ele('@href^tel:', timeout=1)
        if tel_link: telephone = tel_link.attr('href').replace('tel:', '').strip()
        if not telephone: telephone = find_regex(r'(?:Téléphone|Phone)[\s:.]*([\+\d\s\-\(\)]{8,20})')
    except: pass

    adresse = find_regex(r'"streetAddress"\s*:\s*"([^"]+)"')
    if not adresse:
        adr_ele = page.ele('.basic-address-card', timeout=0.5) or page.ele('.contact-address', timeout=0.5)
        if adr_ele: adresse = clean_text(adr_ele.text)

    # Ciblage spécifique du Chiffre d'Affaires dans les tableaux pour éviter les pubs
    ca_brut = ""
    try:
        tr_ca = page.ele('xpath://tr[th[contains(translate(text(), "CHIFFRE D\'AFFAIRES", "chiffre d\'affaires"), "chiffre d\'affaires")]]', timeout=0.5)
        if tr_ca: ca_brut = tr_ca.ele('tag:td').text
    except: pass

    if not ca_brut:
        try:
            texte_visible = page.ele('tag:body').text
            m = re.search(r"(?:Chiffre d'affaires|CA)\s*[:\n]\s*([\d\s,.]+\s*(?:€|M€|K€|MEUR|KEUR|EUR))", texte_visible, re.I)
            if m: ca_brut = m.group(1)
        except: pass

    # Ciblage spécifique des Effectifs
    effectifs_brut = ""
    try:
        tr_eff = page.ele('xpath://tr[th[contains(translate(text(), "EFFECTIFS", "effectifs"), "effectifs")]]', timeout=0.5)
        if tr_eff:
            effectifs_brut = tr_eff.ele('tag:td').text
        else:
            effectifs_brut = find_regex(r"(\d[\d\s]*)\s+(?:salariés?|employés?|collaborateurs?)")
    except: pass

    return {
        "lien_kompass":    meta["lien_kompass"],
        "raison_sociale":  clean_text(meta.get("raison_sociale", "INCONNU")),
        "siret":           siret,
        "siren":           siren,
        "code_naf":        find_regex(r"NAF\s*[:\-]?\s*(\d{4}[A-Z])"),
        "forme_juridique": find_regex(r"\b(SAS|SARL|SA|EURL|SNC|SASU|EI)\b"),
        "date_creation":   find_regex(r"(?:fond[ée]e?|cr[ée][ée]e?)\s+en\s+(\d{4})"),
        "adresse":         adresse[:150], 
        "ville":           clean_text(meta.get("localisation", "").split("-")[0]),
        "code_postal":     find_regex(r'"postalCode"\s*:\s*"(\d{5})"'),
        "site_web":        meta.get("site_web", ""),
        "telephone":       telephone,
        "email":           find_regex(r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})'),
        "chiffre_affaires": nettoyer_chiffre_affaires(ca_brut),
        "effectifs":       nettoyer_effectifs(effectifs_brut),
        "description":     clean_text(meta.get("description", ""))[:250],
        "statut_scraping": "success",
        "date_scraping":   datetime.now().isoformat(),
        "html_brut":       "" 
    }

# ============================================================
# EXPORT EXCEL POWER BI (Format Natif)
# ============================================================
def exporter_excel_propre(conn, chemin_excel):
    colonnes = [
        "id", "lien_kompass", "siret", "siren", "code_naf", "raison_sociale", 
        "forme_juridique", "date_creation", "adresse", "ville", "code_postal", 
        "site_web", "telephone", "email", "chiffre_affaires", "effectifs", "statut_scraping"
    ]
    query = f"SELECT {', '.join(colonnes)} FROM entreprises"
    rows = conn.execute(query).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Base Power BI"

    header_fill = PatternFill(start_color="2A4B7C", end_color="2A4B7C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.append([c.upper().replace('_', ' ') for c in colonnes])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for idx, row in enumerate(rows, start=2):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=idx, column=col_idx, value=val)
            cell.border = border
            
            # Format Text pour les numéros (évite les bugs scientifiques d'Excel)
            if colonnes[col_idx-1] in ["siret", "siren", "code_postal", "telephone"]:
                cell.number_format = '@' 
            # Format Comptabilité pour le CA
            if colonnes[col_idx-1] == "chiffre_affaires" and val is not None:
                cell.number_format = '#,##0 €'
            # Format Entier pour les effectifs
            if colonnes[col_idx-1] == "effectifs" and val is not None:
                cell.number_format = '#,##0'

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = min(max_length + 2, 40)

    wb.save(chemin_excel)
    log.info(f"📊 Fichier Excel généré : {chemin_excel} (Prêt pour Power BI !)")

# ============================================================
# ORCHESTRATEUR PRINCIPAL
# ============================================================
def lancer_test_powerbi():
    try:
        wb = openpyxl.load_workbook(CONFIG["excel_input"], read_only=True)
        ws = wb["Base"]
        toutes_les_urls = [{"lien_kompass": r[0], "raison_sociale": r[1], "localisation": r[2]} for r in list(ws.iter_rows(values_only=True))[1:] if r[0]]
    except Exception as e:
        log.error(f"Erreur de lecture Excel : {e}")
        return

    liens_fr = [u for u in toutes_les_urls if 'fr.kompass.com' in u['lien_kompass']]
    if len(liens_fr) >= CONFIG["taille_test"]:
        echantillon = random.sample(liens_fr, CONFIG["taille_test"])
    else:
        echantillon = random.sample(toutes_les_urls, min(CONFIG["taille_test"], len(toutes_les_urls)))

    base_url = f"{urlparse(echantillon[0]['lien_kompass']).scheme}://{urlparse(echantillon[0]['lien_kompass']).netloc}/"
    conn = init_db(CONFIG["db_path"])
    
    co = ChromiumOptions()
    chemin = trouver_navigateur()
    if chemin: co.set_browser_path(chemin)
    co.headless(False)
    page = ChromiumPage(addr_or_opts=co)
    
    log.info("Warm-up du navigateur furtif...")
    page.get(base_url)
    time.sleep(3)

    for i, meta in enumerate(echantillon, 1):
        url = meta["lien_kompass"]
        print(f"\n--- [{i}/{CONFIG['taille_test']}] Analyse en cours : {meta.get('raison_sociale')} ---")
        
        ok = False
        for tentative in range(CONFIG["max_retries"]):
            try:
                page.get(url)
                page.wait.load_start()
                time.sleep(random.uniform(CONFIG["delay_min"], CONFIG["delay_max"]))
                
                if "kompass" in page.html.lower():
                    try:
                        btn = page.ele('text:Afficher le numéro', timeout=1) or page.ele('.phone-btn', timeout=1)
                        if btn:
                            btn.click()
                            time.sleep(2)
                    except: pass

                    data = extraire(page, meta)
                    sauvegarder_entreprise(conn, data)
                    
                    # Log propre dans la console
                    ca_affiche = f"{data['chiffre_affaires']:,.0f} €".replace(',', ' ') if data['chiffre_affaires'] else 'X'
                    print(f"  ✅ Succès | Tél: {data['telephone'] or 'X'} | CA: {ca_affiche}")
                    
                    ok = True
                    break
            except:
                time.sleep(2)
        
        if not ok:
            print(f"  ❌ Échec de la lecture.")
            sauvegarder_entreprise(conn, {"lien_kompass": url, "statut_scraping": "error", "raison_sociale": meta.get("raison_sociale", "Erreur"), **{k: None for k in ["siret", "siren", "code_naf", "forme_juridique", "date_creation", "adresse", "ville", "code_postal", "site_web", "telephone", "email", "chiffre_affaires", "effectifs", "description", "html_brut"]}, "date_scraping": datetime.now().isoformat()})

    exporter_excel_propre(conn, CONFIG["excel_output"])
    page.quit()
    conn.close()
    
    log.info("\n🚀 TERMINÉ ! Va ouvrir le fichier 'Data_PowerBI_Propre.xlsx'.")

if __name__ == "__main__":
    lancer_test_powerbi()
    