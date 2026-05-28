"""
============================================================
PIPELINE ETL : TRAITEMENT & EXPORT (FINAL)
Synchronisé avec : Scraper Kompass (Version Max Speed)
============================================================
"""
import sqlite3
import pandas as pd
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Configuration
DB_PATH = "base_reindustrialisation.db"
OUTPUT_EXCEL = "Livrable_Final_PowerBI.xlsx"
OUTPUT_JSON = "Livrable_IA_Nettoye.json"

def run_etl():
    print("⚡ Démarrage du pipeline ETL (Traitement)...")
    
    # 1. Vérification de la base de données
    if not os.path.exists(DB_PATH):
        print(f"❌ Erreur : Le fichier '{DB_PATH}' est introuvable. Lance le scraper d'abord.")
        return

    # 2. Extraction SQL
    try:
        conn = sqlite3.connect(DB_PATH)
        df = pd.read_sql("SELECT * FROM entreprises WHERE statut_scraping = 'success'", conn)
        conn.close()
    except Exception as e:
        print(f"❌ Erreur connexion DB : {e}")
        return

    if df.empty:
        print("⚠️ Aucune donnée 'success' trouvée en base.")
        return

    # 3. Nettoyage
    df = df.drop_duplicates(subset=['lien_kompass'])
    df['raison_sociale'] = df['raison_sociale'].astype(str).str.upper().str.strip()
    df['ville'] = df['ville'].astype(str).str.upper().str.strip()

    # 4. Export JSON
    df.to_json(OUTPUT_JSON, orient="records", force_ascii=False, indent=2)
    print(f"✅ Fichier JSON généré : {OUTPUT_JSON}")

    # 5. Export Excel avec Sécurité Anti-Crash
    fichier_excel_final = OUTPUT_EXCEL
    try:
        df.to_excel(fichier_excel_final, index=False)
    except PermissionError:
        print(f"\n⚠️ ATTENTION : Le fichier '{OUTPUT_EXCEL}' est ouvert dans Excel !")
        fichier_excel_final = "Livrable_Final_PowerBI_V2.xlsx"
        print(f"🔄 Sauvegarde de secours activée sous le nom : {fichier_excel_final}\n")
        df.to_excel(fichier_excel_final, index=False)
    
    # 6. Stylisation Excel
    wb = load_workbook(fichier_excel_final)
    ws = wb.active
    
    header_fill = PatternFill(start_color="2A4B7C", end_color="2A4B7C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            # Format texte pour SIRET/SIREN/CP/Tel/Email
            if cell.column in [3, 4, 10, 11, 12]: 
                cell.number_format = '@' 
            # Format Monétaire pour le CA
            if cell.column == 13 and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0 €'
    
    wb.save(fichier_excel_final)
    print(f"✅ Fichier Excel formaté généré : {fichier_excel_final}")
    print(f"📊 Traitement terminé : {len(df)} entreprises exportées avec succès.")

if __name__ == "__main__":
    run_etl()