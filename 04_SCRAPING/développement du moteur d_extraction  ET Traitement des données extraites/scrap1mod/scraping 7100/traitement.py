"""
PIPELINE ETL : TRAITEMENT & EXPORT (FINAL)
Synchronisé avec la base SQL du scraper.
Génère : 
1. Livrable_Final_PowerBI.xlsx (Formaté)
2. Livrable_IA_Nettoye.json (Structuré)
"""
import sqlite3
import pandas as pd
import json
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Configuration
DB_PATH = "base_reindustrialisation_test.db"
OUTPUT_EXCEL = "Livrable_Final_PowerBI.xlsx"
OUTPUT_JSON = "Livrable_IA_Nettoye.json"

def run_etl():
    print("⚡ Démarrage du pipeline ETL...")
    
    # 1. Extraction SQL via Pandas
    try:
        conn = sqlite3.connect(DB_PATH)
        # On ne récupère que les lignes réussies
        df = pd.read_sql("SELECT * FROM entreprises WHERE statut_scraping = 'success'", conn)
        conn.close()
    except Exception as e:
        print(f"❌ Erreur connexion DB : {e}")
        return

    if df.empty:
        print("⚠️ Aucune donnée trouvée. Vérifiez que le scraper a bien rempli la base.")
        return

    # 2. Nettoyage
    df = df.drop_duplicates(subset=['lien_kompass'])
    df['raison_sociale'] = df['raison_sociale'].str.upper().str.strip()
    df['ville'] = df['ville'].str.upper().str.strip()

    # 3. Export JSON (Pour l'IA)
    # L'export JSON se fait avant le formatage Excel car il a besoin de données brutes
    df.to_json(OUTPUT_JSON, orient="records", force_ascii=False, indent=2)
    print(f"✅ JSON généré : {OUTPUT_JSON}")

    # 4. Export Excel Natif
    df.to_excel(OUTPUT_EXCEL, index=False)
    
    # 5. Stylisation Pro pour Power BI
    wb = load_workbook(OUTPUT_EXCEL)
    ws = wb.active
    
    # Style en-têtes
    header_fill = PatternFill(start_color="2A4B7C", end_color="2A4B7C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Style données
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            # Format SIREN/SIRET/Tél en texte (colonne 3=siret, 4=siren, 11=tel, 12=email)
            if cell.column in [3, 4, 11, 12]:
                cell.number_format = '@'
            # Format CA (colonne 13)
            if cell.column == 13 and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0 €'
    
    wb.save(OUTPUT_EXCEL)
    print(f"✅ Excel formaté généré : {OUTPUT_EXCEL}")
    print(f"📊 Traitement terminé : {len(df)} entreprises traitées.")

if __name__ == "__main__":
    run_etl()