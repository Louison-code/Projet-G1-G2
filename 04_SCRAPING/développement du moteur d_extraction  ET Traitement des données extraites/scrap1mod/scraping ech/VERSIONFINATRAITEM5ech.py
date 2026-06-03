"""
============================================================
PROJET G1/G2 - F-2049 | Réindustrialisation Française
T3.1.4 : Traitement des données extraites (Version Power BI)
============================================================
Auteurs (RACI R) : Yessine Hachicha
Accountable (A)  : Louison Baudouin
------------------------------------------------------------
Pipeline ETL officiel :
  1. Extraction depuis SQLite (base_reindustrialisation_test.db)
  2. Transformation (Standardisation texte, gestion des Nulls)
  3. Chargement : Export EXCEL formaté pour Power BI (T3)
  4. Chargement : Export JSON structuré pour l'IA (T2)
============================================================
"""

import sqlite3
import re
import json
import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURATION
# ============================================================
CONFIG = {
    "db_path":           "base_reindustrialisation_test.db",
    "export_excel":      "Livrable_PowerBI_Nettoye.xlsx", # Remplacement du CSV
    "export_json":       "Livrable_IA_Nettoye.json",
    "log_path":          "traitement_donnees.log",
    "seuil_ca_min":      0,              
    "seuil_effectifs_min": 0,
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(CONFIG["log_path"], encoding="utf-8"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("TraitementData")

# ============================================================
# TRANSFORMATION & NETTOYAGE
# ============================================================
def nettoyer_texte(texte):
    """Supprime les espaces multiples et sauts de ligne."""
    if not texte: return ""
    return re.sub(r'\s+', ' ', str(texte)).strip()

def pipeline_nettoyage(db_path):
    """Extrait, nettoie et filtre les données de la base SQL."""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    
    rows = conn.execute("SELECT * FROM entreprises WHERE statut_scraping = 'success'").fetchall()
    conn.close()

    entreprises_propres = []
    doublons_siren = set()

    for r in rows:
        ent = dict(r)
        
        # 1. Standardisation du texte
        ent["raison_sociale"] = nettoyer_texte(ent.get("raison_sociale")).upper()
        ent["ville"] = nettoyer_texte(ent.get("ville")).upper()
        
        # 2. Correction sécurisée pour les types numériques
        # On extrait la valeur, si c'est vide ou None, on force à 0
        ca_raw = ent.get("chiffre_affaires")
        effectifs_raw = ent.get("effectifs")
        
        # Conversion sécurisée : on gère les chaînes vides, les None et les espaces
        try:
            ca = float(ca_raw) if (ca_raw is not None and str(ca_raw).strip() != "") else 0.0
        except ValueError:
            ca = 0.0
            
        try:
            effectifs = int(float(effectifs_raw)) if (effectifs_raw is not None and str(effectifs_raw).strip() != "") else 0
        except ValueError:
            effectifs = 0

        # Filtres métiers
        if ca < CONFIG["seuil_ca_min"] or effectifs < CONFIG["seuil_effectifs_min"]:
            continue

        ent["chiffre_affaires"] = ca if ca > 0 else None
        ent["effectifs"] = effectifs if effectifs > 0 else None

        # 3. Dédoublonnage
        siren = ent.get("siren")
        if siren:
            if siren in doublons_siren: continue
            doublons_siren.add(siren)

        if "html_brut" in ent: del ent["html_brut"]
        entreprises_propres.append(ent)

    logger.info(f"Nettoyage terminé : {len(entreprises_propres)} entreprises validées.")
    return entreprises_propres

# ============================================================
# EXPORT EXCEL (Spécial Power BI)
# ============================================================
def exporter_excel(entreprises, chemin):
    """Génère un fichier Excel natif, formaté et prêt pour Power BI."""
    if not entreprises: return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Base Nettoyée"

    colonnes = list(entreprises[0].keys())
    
    # Styles
    header_fill = PatternFill(start_color="2A4B7C", end_color="2A4B7C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Écriture des en-têtes
    ws.append([c.upper().replace('_', ' ') for c in colonnes])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Écriture des données
    for idx, ent in enumerate(entreprises, start=2):
        for col_idx, key in enumerate(colonnes, start=1):
            val = ent[key]
            cell = ws.cell(row=idx, column=col_idx, value=val)
            cell.border = border
            
            # Formatage spécifique
            if key in ["siret", "siren", "code_postal", "telephone"]:
                cell.number_format = '@' # Texte pour garder les zéros
                cell.alignment = Alignment(horizontal="center")
            elif key == "chiffre_affaires" and val is not None:
                cell.number_format = '#,##0 €'
            elif key == "effectifs" and val is not None:
                cell.number_format = '#,##0'

    # Ajustement des largeurs de colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = min(max_length + 2, 40)

    wb.save(chemin)
    logger.info(f"Export Excel Power BI : {chemin}")

# ============================================================
# EXPORT JSON (Pour le Groupe IA)
# ============================================================
def exporter_json(entreprises, chemin):
    """Export JSON structuré pour le Groupe IA (T2)."""
    with open(chemin, "w", encoding="utf-8") as f:
        json.dump(entreprises, f, ensure_ascii=False, indent=2, default=str)
    logger.info(f"Export JSON IA : {chemin}")

# ============================================================
# POINT D'ENTRÉE
# ============================================================
def main():
    logger.info("=" * 60)
    logger.info("PROJET G1/G2 | Pipeline Traitement Données (ETL)")
    logger.info(f"Démarrage : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("=" * 60)

    db_path = CONFIG["db_path"]
    if not Path(db_path).exists():
        logger.error(f"Base de données introuvable : {db_path}")
        logger.error("Lancez d'abord le Scraper V8 pour générer la base.")
        return

    # Pipeline de nettoyage
    entreprises_propres = pipeline_nettoyage(db_path)

    if not entreprises_propres:
        logger.warning("Aucune donnée propre disponible. Vérifiez la base.")
        return

    # Exports
    exporter_excel(entreprises_propres, CONFIG["export_excel"])
    exporter_json(entreprises_propres,  CONFIG["export_json"])

    logger.info("=" * 60)
    logger.info("TRAITEMENT TERMINÉ AVEC SUCCÈS")
    logger.info("=" * 60)

if __name__ == "__main__":
    main()
    