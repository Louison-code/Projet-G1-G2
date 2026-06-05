#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Pipeline Post-Traitement
========================
Nettoie, enrichit et exporte les donnees scrapées.
Execute APRES les scrapers pour preparer les livrables.

Usage:
    python pipeline_post_traitement.py          # Tout faire
    python pipeline_post_traitement.py --nettoyage  # Nettoyage seul
    python pipeline_post_traitement.py --export      # Export seul
"""

import re, json, csv, os, sys, sqlite3, time
from datetime import datetime
from pathlib import Path

BASE_DIR = Path(__file__).parent.resolve()
DB_PATH = BASE_DIR / "data" / "base_reindustrialisation.db"
EXPORT_DIR = BASE_DIR / "exports"
EXPORT_DIR.mkdir(exist_ok=True)

COLONNES_EXPORT = [
    "url", "nom_entreprise", "roles", "description", "code_postal",
    "ville", "pays", "telephone", "fax", "email", "site_web",
    "siren", "siret", "tva", "capital", "forme_juridique",
    "annee_creation", "effectif_adresse", "effectif_entreprise",
    "activites_principales", "activites_secondaires", "autres_classifications",
    "code_naf", "departement", "region", "chiffre_affaires",
    "secteur_ia", "filiere_ia", "latitude", "longitude",
]


# ───────────────────────────────────────────
# 1. NETTOYAGE
# ───────────────────────────────────────────

def nettoyer():
    print("\n=== 1. NETTOYAGE ===")
    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()
    stats = {}

    # 1.1 Supprimer les lignes sans SIREN (inutilisables)
    cur.execute("DELETE FROM entreprises WHERE siren IS NULL OR TRIM(siren) = ''")
    stats["sans_siren"] = cur.rowcount
    print(f"  Lignes sans SIREN supprimees : {stats['sans_siren']}")

    # 1.2 Doublons de SIRET (garder la plus complete)
    cur.execute("""
        DELETE FROM entreprises WHERE id IN (
            SELECT id FROM (
                SELECT id, siret,
                    ROW_NUMBER() OVER (
                        PARTITION BY siret ORDER BY
                            (LENGTH(COALESCE(nom_entreprise,'')) +
                             LENGTH(COALESCE(telephone,'')) +
                             LENGTH(COALESCE(email,'')) +
                             LENGTH(COALESCE(activites_principales,''))) DESC
                    ) AS rn
                FROM entreprises
                WHERE siret IS NOT NULL AND TRIM(siret) != ''
            ) WHERE rn > 1
        )
    """)
    stats["doublons_siret"] = cur.rowcount
    print(f"  Doublons SIRET supprimes : {stats['doublons_siret']}")

    # 1.3 Doublons de SIREN
    cur.execute("""
        DELETE FROM entreprises WHERE id IN (
            SELECT id FROM (
                SELECT id, siren,
                    ROW_NUMBER() OVER (
                        PARTITION BY siren ORDER BY
                            (LENGTH(COALESCE(nom_entreprise,'')) +
                             LENGTH(COALESCE(telephone,'')) +
                             LENGTH(COALESCE(email,'')) +
                             LENGTH(COALESCE(activites_principales,''))) DESC
                    ) AS rn
                FROM entreprises
                WHERE siren IS NOT NULL AND TRIM(siren) != ''
            ) WHERE rn > 1
        )
    """)
    stats["doublons_siren"] = cur.rowcount
    print(f"  Doublons SIREN supprimes : {stats['doublons_siren']}")

    # 1.4 Normaliser les noms de ville
    cur.execute("SELECT id, ville FROM entreprises WHERE ville IS NOT NULL AND ville != ''")
    ids_ville = 0
    for row_id, ville in cur.fetchall():
        ville_propre = re.sub(r'\s+', ' ', ville).strip().upper()
        if ville != ville_propre:
            conn.execute("UPDATE entreprises SET ville = ? WHERE id = ?", (ville_propre, row_id))
            ids_ville += 1
    print(f"  Villes normalisees (majuscules) : {ids_ville}")

    # 1.5 Normaliser les emails (minuscules)
    cur.execute("SELECT id, email FROM entreprises WHERE email IS NOT NULL AND email != ''")
    ids_email = 0
    for row_id, email in cur.fetchall():
        email_propre = email.strip().lower()
        if email != email_propre:
            conn.execute("UPDATE entreprises SET email = ? WHERE id = ?", (email_propre, row_id))
            ids_email += 1
    print(f"  Emails normalises : {ids_email}")

    # 1.6 Normaliser les SIREN (9 chiffres)
    cur.execute("SELECT id, siren FROM entreprises WHERE siren IS NOT NULL AND siren != ''")
    ids_siren = 0
    for row_id, siren in cur.fetchall():
        siren_propre = re.sub(r"\D", "", siren)
        if len(siren_propre) != 9:
            conn.execute("UPDATE entreprises SET siren = NULL WHERE id = ?", (row_id,))
        elif siren_propre != siren:
            conn.execute("UPDATE entreprises SET siren = ? WHERE id = ?", (siren_propre, row_id))
            ids_siren += 1
    print(f"  SIREN normalises : {ids_siren}")

    conn.commit()

    # Statistiques finales
    cur.execute("SELECT COUNT(*) FROM entreprises")
    total = cur.fetchone()[0]
    print(f"  Total entreprises apres nettoyage : {total}")

    conn.close()
    return total


# ───────────────────────────────────────────
# 2. INDICATEURS
# ───────────────────────────────────────────

def calculer_indicateurs():
    print("\n=== 2. INDICATEURS ===")
    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()

    cur.execute("""
        SELECT id, chiffre_affaires, effectif_entreprise, annee_creation
        FROM entreprises
        WHERE chiffre_affaires IS NOT NULL OR effectif_entreprise IS NOT NULL
    """)

    compteur = 0
    for row_id, ca, effectif, annee in cur.fetchall():
        annee_val = None
        if annee and re.match(r'^\d{4}$', str(annee).strip()):
            annee_val = int(annee.strip())
        elif annee:
            m = re.search(r'\b(19[3-9]\d|20[0-2]\d)\b', str(annee))
            if m:
                annee_val = int(m.group(1))

        if annee_val is None:
            annee_val = datetime.now().year

        effectif_int = None
        if effectif:
            eff_clean = re.sub(r"\D", "", str(effectif))
            if eff_clean:
                try:
                    effectif_int = int(eff_clean)
                except:
                    pass

        ca_float = None
        if ca:
            try:
                ca_float = float(re.sub(r"[^\d.,]", "", str(ca).replace(" ", "")).replace(",", "."))
            except:
                pass

        if ca_float is not None or effectif_int is not None:
            conn.execute(
                "INSERT OR IGNORE INTO indicateurs (entreprise_id, annee, chiffre_affaires, effectifs) "
                "VALUES (?, ?, ?, ?)",
                (row_id, annee_val, ca_float, effectif_int)
            )
            compteur += 1

    conn.commit()
    conn.close()
    print(f"  Indicateurs crees : {compteur}")


# ───────────────────────────────────────────
# 3. GEOLOCALISATION
# ───────────────────────────────────────────

def geolocaliser():
    print("\n=== 3. GEOLOCALISATION ===")
    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()

    cur.execute("""
        SELECT id, code_postal, ville, adresse
        FROM entreprises
        WHERE (latitude IS NULL OR longitude IS NULL)
          AND code_postal IS NOT NULL AND TRIM(code_postal) != ''
          AND ville IS NOT NULL AND TRIM(ville) != ''
        LIMIT 500
    """)

    entreprises = cur.fetchall()
    if not entreprises:
        print("  Aucune entreprise a geolocaliser (deja fait ou pas d'adresse)")
        conn.close()
        return

    total = len(entreprises)
    reussi = 0
    echecs = 0

    try:
        import requests
        for row_id, cp, ville, adresse in entreprises:
            try:
                query = f"{ville} {cp} France"
                if adresse and len(adresse) > 5:
                    query = f"{adresse}, {cp} {ville} France"

                r = requests.get(
                    "https://nominatim.openstreetmap.org/search",
                    params={"q": query, "format": "json", "limit": 1, "countrycodes": "fr"},
                    headers={"User-Agent": "PipelinePostTraitement/1.0"},
                    timeout=10
                )

                if r.status_code == 200 and r.json():
                    data = r.json()[0]
                    lat = float(data.get("lat", 0))
                    lng = float(data.get("lon", 0))
                    conn.execute(
                        "UPDATE entreprises SET latitude = ?, longitude = ? WHERE id = ?",
                        (lat, lng, row_id)
                    )
                    reussi += 1
                else:
                    echecs += 1

                time.sleep(1.1)

            except Exception:
                echecs += 1

    except ImportError:
        print("  requests non installe. Installation: pip install requests")
        conn.close()
        return

    conn.commit()
    conn.close()
    print(f"  Geolocalisees : {reussi}/{total} (echecs: {echecs})")


# ───────────────────────────────────────────
# 4. CLASSIFICATION IA / FILIERE
# ───────────────────────────────────────────

MOTS_CLES_IA = {
    "fabrication additive": ("Fabrication additive", "Production"),
    "impression 3d": ("Fabrication additive", "Production"),
    "robotique": ("Robotique", "Production"),
    "automatisation": ("Robotique", "Production"),
    "ia": ("Intelligence artificielle", "Conception"),
    "intelligence artificielle": ("Intelligence artificielle", "Conception"),
    "machine learning": ("Intelligence artificielle", "Conception"),
    "deep learning": ("Intelligence artificielle", "Conception"),
    "vision artificielle": ("Vision industrielle", "Controle qualite"),
    "vision industrielle": ("Vision industrielle", "Controle qualite"),
    "traitement image": ("Vision industrielle", "Controle qualite"),
    "iot": ("IoT", "Production"),
    "internet des objets": ("IoT", "Production"),
    "capteur": ("IoT", "Production"),
    "big data": ("Big Data", "Analyse"),
    "data science": ("Big Data", "Analyse"),
    "analyse donnee": ("Big Data", "Analyse"),
    "cloud": ("Cloud computing", "Infrastructure"),
    "cybersecurite": ("Cybersecurite", "Infrastructure"),
    "maintenance predictive": ("Maintenance predictive", "Production"),
    "maintenance previsionnelle": ("Maintenance predictive", "Production"),
    "jumeau numerique": ("Jumeau numerique", "Conception"),
    "digital twin": ("Jumeau numerique", "Conception"),
    "simulation": ("Simulation", "Conception"),
    "cobotique": ("Cobotique", "Production"),
    "collaboratif": ("Cobotique", "Production"),
    "energie": ("Efficacite energetique", "Environnement"),
    "decarbonation": ("Decarbonation", "Environnement"),
    "decarbone": ("Decarbonation", "Environnement"),
}

FILIERE_PAR_NAF = {
    "28": "Machines et equipements",
    "29": "Industrie automobile",
    "30": "Equipements de transport",
    "25": "Metallurgie",
    "26": "Electronique",
    "27": "Equipements electriques",
    "20": "Chimie",
    "21": "Pharmacie",
    "22": "Plasturgie",
    "23": "Materiaux de construction",
    "24": "Metallurgie",
    "31": "Meubles",
    "32": "Autres industries",
    "33": "Maintenance industrielle",
    "10": "Agroalimentaire",
    "11": "Agroalimentaire",
    "12": "Agroalimentaire",
}


def classer_entreprises():
    print("\n=== 4. CLASSIFICATION IA/FILIERE ===")
    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()

    cur.execute("""
        SELECT id, activites_principales, description, code_naf
        FROM entreprises
    """)

    compteur = 0
    for row_id, ap, desc, naf in cur.fetchall():
        texte = f"{ap or ''} {desc or ''} {naf or ''}".lower()

        secteur = ""
        filiere = ""

        # Detection par mots-cles IA
        meilleur_score = 0
        for mot_cle, (sect, fil) in MOTS_CLES_IA.items():
            if mot_cle in texte:
                score = len(mot_cle) / max(len(texte), 1)
                if score > meilleur_score:
                    meilleur_score = score
                    secteur = sect
                    filiere = fil

        # Fallback : filiere par code NAF
        if naf:
            naf_prefix = naf[:2] if naf else ""
            if naf_prefix in FILIERE_PAR_NAF:
                filiere = filiere or FILIERE_PAR_NAF[naf_prefix]

        secteur = secteur or "Non classe"
        filiere = filiere or "Non classee"

        if secteur or filiere:
            conn.execute(
                "UPDATE entreprises SET secteur_ia = ?, filiere_ia = ? WHERE id = ?",
                (secteur, filiere, row_id)
            )
            compteur += 1

    conn.commit()
    conn.close()
    print(f"  Entreprises classees : {compteur}")


# ───────────────────────────────────────────
# 5. EXPORT
# ───────────────────────────────────────────

def exporter():
    print("\n=== 5. EXPORT ===")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()

    colonnes_sql = ", ".join(f'"{c}"' for c in COLONNES_EXPORT)
    cur.execute(f"SELECT {colonnes_sql} FROM entreprises ORDER BY nom_entreprise")
    rows = cur.fetchall()
    conn.close()

    total = len(rows)
    if total == 0:
        print("  Aucune donnee a exporter.")
        return

    # ── Excel ──
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        xlsx_path = EXPORT_DIR / f"Livrable_PowerBI_{ts}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Entreprises"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2F5496")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for ci, col in enumerate(COLONNES_EXPORT, 1):
            cell = ws.cell(1, ci, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(ri, ci, val)
                cell.border = thin_border

        ws.auto_filter.ref = f"A1:{chr(64+len(COLONNES_EXPORT))}{total+1}"
        ws.freeze_panes = "A2"

        for ci, col in enumerate(COLONNES_EXPORT, 1):
            max_len = len(str(col))
            for ri in range(2, min(total + 2, 50)):
                cell_val = ws.cell(ri, ci).value
                if cell_val:
                    max_len = max(max_len, min(len(str(cell_val)), 50))
            ws.column_dimensions[chr(64+ci) if ci < 27 else 'A'* (ci//26) + chr(64+ci%26)].width = max_len + 2

        wb.save(str(xlsx_path))
        print(f"  Excel : {xlsx_path} (lignes: {total})")

    except ImportError:
        print(f"  Excel : openpyxl non installe (pip install openpyxl)")

    # ── CSV ──
    csv_path = EXPORT_DIR / f"Livrable_CSV_{ts}.csv"
    with open(str(csv_path), "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(COLONNES_EXPORT)
        for row in rows:
            w.writerow([str(v) if v else "" for v in row])
    print(f"  CSV  : {csv_path} (lignes: {total})")

    # ── JSON ──
    json_path = EXPORT_DIR / f"Livrable_IA_{ts}.json"
    data_export = []
    for row in rows:
        data_export.append(dict(zip(COLONNES_EXPORT, [str(v) if v else "" for v in row])))
    with open(str(json_path), "w", encoding="utf-8") as f:
        json.dump(data_export, f, ensure_ascii=False, indent=2)
    print(f"  JSON : {json_path} (lignes: {total})")


# ───────────────────────────────────────────
# RAPPORT DE QUALITE
# ───────────────────────────────────────────

def rapport_qualite(total_apres_nettoyage: int = None):
    print("\n=== RAPPORT DE QUALITE ===")
    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM entreprises")
    total = cur.fetchone()[0]
    print(f"  Total entreprises : {total}")

    for col in COLONNES_EXPORT[:22]:
        cur.execute(f"SELECT COUNT(*) FROM entreprises WHERE {col} IS NOT NULL AND TRIM({col}) != ''")
        remplis = cur.fetchone()[0]
        pct = remplis * 100 // max(total, 1)
        barre = "#" * (remplis * 30 // max(total, 1))
        print(f"    {col:30s} {remplis:4d}/{total} ({pct:3d}%) {barre}")

    conn.close()


# ───────────────────────────────────────────
# MAIN
# ───────────────────────────────────────────

def main():
    args = sys.argv[1:] if len(sys.argv) > 1 else ["--tout"]

    do_nettoyage = "--tout" in args or "--nettoyage" in args
    do_indicateurs = "--tout" in args or "--indicateurs" in args
    do_geo = "--tout" in args or "--geolocalisation" in args
    do_classification = "--tout" in args or "--classification" in args
    do_export = "--tout" in args or "--export" in args

    print("=" * 60)
    print("  PIPELINE POST-TRAITEMENT")
    print(f"  Base : {DB_PATH}")
    print(f"  Date : {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("=" * 60)

    # Verification que la base existe
    if not DB_PATH.exists():
        print(f"\n  ERREUR : Base introuvable -> {DB_PATH}")
        sys.exit(1)

    total_apres = None

    if do_nettoyage:
        total_apres = nettoyer()

    if do_indicateurs:
        calculer_indicateurs()

    if do_geo:
        geolocaliser()

    if do_classification:
        classer_entreprises()

    if do_export:
        exporter()

    rapport_qualite(total_apres)

    print("\n  Pipeline termine.\n")


if __name__ == "__main__":
    main()
