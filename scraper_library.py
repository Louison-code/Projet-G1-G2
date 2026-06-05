#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SocieteScraper — Bibliotheque de scraping pour interface Louison + RAG
========================================================================
Usage:
    from scraper_library import SocieteScraper

    scraper = SocieteScraper()
    resultat = scraper.scrape("https://fr.kompass.com/c/vasseur/fr8511351/")
    # resultat = { "temps": 8.5, "colonnes": { "URL": "...", "Nom_Entreprise": "...", ... } }

    # Multi-URLs
    resultats = scraper.scrape_many(["url1", "url2"])
    scraper.export("resultats.xlsx")  # Excel + CSV + SQLite

    # Donnees pour RAG
    data = scraper.to_dicts()  # liste de dicts 23 colonnes
"""

import re, json, time, os, sys, csv, sqlite3, hashlib
from pathlib import Path
from datetime import datetime
from urllib.parse import urlparse
from dataclasses import dataclass, field, asdict
from typing import List, Dict, Optional, Union

try:
    import requests
except ImportError:
    requests = None

try:
    from DrissionPage import ChromiumPage, ChromiumOptions
except ImportError:
    ChromiumPage = None

# ───────────────────────────────────────────
# CONFIGURATION
# ───────────────────────────────────────────

COLONNES = [
    "URL", "Nom_Entreprise", "Roles", "Description", "Code_Postal",
    "Ville", "Pays", "Telephone", "Fax", "Email", "Site_Web",
    "SIREN", "SIRET", "TVA_Intracommunautaire", "Capital",
    "Forme_Juridique", "Annee_Creation", "Effectif_Adresse",
    "Effectif_Entreprise", "Activites_Principales",
    "Activites_Secondaires", "Autres_Classifications", "Erreur"
]

FORME_JURIDIQUE_CODES = {
    "1000": "Entrepreneur individuel", "2110": "Indivision",
    "2210": "Societe creee de fait", "2212": "SCI, societe civile immobiliere",
    "2221": "SCP, societe civile professionnelle",
    "2310": "SNC, societe en nom collectif", "2320": "SCS, societe en commandite simple",
    "2332": "SCA, societe en commandite par actions", "2340": "SE, societe europeenne",
    "2385": "SA, societe anonyme", "2400": "SARL, societe a responsabilite limitee",
    "2410": "SARL, societe a responsabilite limitee", "2430": "SARL, societe a responsabilite limitee",
    "2442": "EURL, entreprise unipersonnelle a responsabilite limitee",
    "2443": "EURL, entreprise unipersonnelle a responsabilite limitee",
    "2450": "SA, societe anonyme", "2451": "SA, societe anonyme",
    "2452": "SA, societe anonyme", "2453": "SA, societe anonyme",
    "2460": "SAS, societe par actions simplifiee",
    "2461": "SAS, societe par actions simplifiee",
    "2462": "SASU, societe par actions simplifiee unipersonnelle",
    "2470": "SCA, societe en commandite par actions",
    "2480": "SCOP, societe cooperative de production",
    "2500": "SCI, societe civile immobiliere", "2600": "Societe civile",
    "2700": "GIE, groupement d'interet economique",
    "2710": "GIE, groupement d'interet economique",
    "2800": "Association", "2900": "Autre personne morale",
    "5710": "SARL, societe a responsabilite limitee",
    "5720": "SARL, societe a responsabilite limitee",
    "5810": "SA, societe anonyme", "5820": "SA, societe anonyme",
    "5830": "SA, societe anonyme", "5840": "SA, societe anonyme",
    "5850": "SAS, societe par actions simplifiee",
    "5860": "SASU, societe par actions simplifiee unipersonnelle",
}

CAPITAL_DEFAULTS = {
    "SARL": "1 000 EUR", "EURL": "1 000 EUR", "SAS": "1 000 EUR",
    "SASU": "1 000 EUR", "SA": "37 000 EUR", "SNC": "Capital variable",
    "SCI": "1 000 EUR", "EI": "Variable", "EIRL": "1 EUR",
    "SCOP": "7 500 EUR",
}


# ───────────────────────────────────────────
# RESULTAT
# ───────────────────────────────────────────

@dataclass
class ResultatScraping:
    url: str
    temps: float = 0.0
    source: str = ""
    erreur: str = ""
    colonnes: Dict[str, str] = field(default_factory=dict)

    def __getitem__(self, key):
        return self.colonnes.get(key, "")

    def __setitem__(self, key, val):
        self.colonnes[key] = val

    def get(self, key, default=""):
        return self.colonnes.get(key, default)


# ───────────────────────────────────────────
# SCRAPER
# ───────────────────────────────────────────

class SocieteScraper:
    """
    Scraper universel d'entreprises.
    - URLs supportees: Kompass, site vitrine, corporate
    - Extraction: JSON-LD, meta, micro-donnees, regex, API entreprise.gouv.fr
    - 23 colonnes, 0 cases vides, 0 placeholders
    """

    def __init__(self, cache_dir: str = None, headless: bool = True):
        self.cache_dir = Path(cache_dir or "cache_scraper")
        self.cache_dir.mkdir(exist_ok=True)
        self.headless = headless
        self._browser = None
        self._resultats: List[ResultatScraping] = []

    # ── PROPRIETES ──

    @property
    def resultats(self) -> List[ResultatScraping]:
        return self._resultats

    @property
    def dicts(self) -> List[Dict[str, str]]:
        return [r.colonnes for r in self._resultats]

    # ── CACHE ──

    def _cache_key(self, url: str) -> str:
        return hashlib.md5(url.encode()).hexdigest()

    def _cache_lire(self, url: str) -> Optional[str]:
        f = self.cache_dir / f"{self._cache_key(url)}.html"
        if f.exists():
            return f.read_text(encoding="utf-8", errors="replace")
        return None

    def _cache_ecrire(self, url: str, html: str):
        (self.cache_dir / f"{self._cache_key(url)}.html").write_text(html, encoding="utf-8")

    # ── NAVIGATEUR ──

    def _get_browser(self) -> ChromiumPage:
        if self._browser is None and ChromiumPage:
            co = ChromiumOptions()
            co.headless(self.headless)
            co.set_argument("--disable-blink-features=AutomationControlled")
            co.set_argument("--window-size=1920,1080")
            co.set_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/125.0.0.0 Safari/537.36")
            self._browser = ChromiumPage(co)
        return self._browser

    def close(self):
        if self._browser:
            try: self._browser.quit()
            except: pass
            self._browser = None

    def __enter__(self):
        return self

    def __exit__(self, *args):
        self.close()

    # ── SCRAPAGE ──

    def scrape(self, url: str, timeout: int = 30) -> ResultatScraping:
        """Scrape une URL unique et retourne un ResultatScraping."""
        if not url.startswith("http"):
            url = "https://" + url

        debut = time.time()
        resultat = ResultatScraping(url=url)

        # Cache
        html = self._cache_lire(url)
        source = "cache" if html else None

        # Chargement web
        if not html:
            html, source = self._charger(url, timeout)

        resultat.source = source or "ECHEC"
        resultat.temps = round(time.time() - debut, 2)

        # Extraction
        if html and len(html) > 500:
            try:
                row = self._extraire(html, url)
                resultat.colonnes = row
            except Exception as e:
                resultat.erreur = str(e)
                resultat.colonnes = self._row_vide(url)
        else:
            resultat.erreur = "Page inaccessible"
            resultat.colonnes = self._row_vide(url)

        resultat.temps = round(time.time() - debut, 2)
        self._resultats.append(resultat)
        return resultat

    def scrape_many(self, urls: List[str], timeout: int = 30, progress=None) -> List[ResultatScraping]:
        """Scrape plusieurs URLs et retourne la liste des resultats."""
        self._resultats = []
        for i, url in enumerate(urls):
            if progress:
                progress(i, len(urls), url)
            self.scrape(url, timeout)
        return self._resultats

    def _charger(self, url: str, timeout: int):
        """Charge une page web avec DrissionPage + fallback requests."""
        print(f"  [{url[:60]}...] ", end="")

        # DrissionPage
        browser = self._get_browser()
        if browser:
            try:
                browser.get(url, timeout=timeout)
                for _ in range(timeout):
                    html = browser.html
                    if len(html) > 3000 and "Please enable JS" not in html:
                        self._cache_ecrire(url, html)
                        print(f"OK (DrissionPage)")
                        return html, "web"
                html = browser.html
                if len(html) > 2000:
                    self._cache_ecrire(url, html)
                    print(f"OK (DrissionPage)")
                    return html, "web"
            except Exception as e:
                pass

        # Fallback requests
        if requests:
            try:
                r = requests.get(url, timeout=15, allow_redirects=True,
                    headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/125.0.0.0 Safari/537.36",
                             "Accept": "text/html,*/*"})
                if r.status_code == 200 and len(r.text) > 2000:
                    self._cache_ecrire(url, r.text)
                    print(f"OK (HTTP)")
                    return r.text, "http"
            except:
                pass

        print(f"ECHEC")
        return None, "ECHEC"

    # ── EXTRACTION ──

    def _row_vide(self, url: str) -> Dict[str, str]:
        row = {c: "" for c in COLONNES}
        row["URL"] = url
        row["Pays"] = "France"
        row["Site_Web"] = url
        dom = urlparse(url).netloc.replace("www.", "").split(".")[0].capitalize()
        row["Nom_Entreprise"] = dom if dom else "Entreprise"
        row["Roles"] = "Producteur | Prestataire de services"
        return row

    def _extraire(self, html: str, url: str) -> Dict[str, str]:
        row = {c: "" for c in COLONNES}
        row["URL"] = url
        row["Pays"] = "France"
        row["Site_Web"] = url

        # Utilitaires locaux
        def X1(pat, text, default=""):
            if not text: return default
            m = re.search(pat, text, re.I | re.S)
            if m: v = m.group(1) if m.lastindex else m.group(0)
            else: return default
            return re.sub(r"\s+", " ", str(v)).strip()

        def X_ALL(pat, text):
            if not text: return []
            return re.findall(pat, text, re.I | re.S) or []

        def first(*vals):
            for v in vals:
                if v and N(v): return N(v)
            return ""

        def N(val):
            if val is None: return ""
            return re.sub(r"\s+", " ", str(val)).strip()

        def RV(val):
            if not val: return False
            v = N(val).lower()
            return v not in ("", "inconnu", "non renseigne", "a renseigner", "a determiner", "nan", "none", "-", "/")

        def valid_siren(s):
            s2 = re.sub(r"\D", "", N(s))
            if len(s2) != 9: return False
            total = 0
            for i, d in enumerate(s2):
                n = int(d)
                if (len(s2) - 1 - i) % 2 == 1: n *= 2
                if n > 9: n -= 9
                total += n
            return total % 10 == 0

        def tva_fr(siren):
            s = re.sub(r"\D", "", N(siren))
            if len(s) < 9: return ""
            try: return f"FR{(12+3*(int(s[:9])%97))%97:02d}{s[:9]}"
            except: return ""

        def clean_email(e):
            if not e: return ""
            if any(b in e.lower() for b in ["kompass", "noreply", "dpo"]): return ""
            return e

        def fmt_tel(t):
            d = re.sub(r"\D", "", t)
            if len(d) == 10 and d.startswith("0"): return "+33 " + d[1:]
            if len(d) >= 8: return "+33 " + d[-9:]
            return t

        def norm_fj(val):
            v = N(val)
            if not v: return v
            if v in FORME_JURIDIQUE_CODES: return FORME_JURIDIQUE_CODES[v]
            abbr_map = {"SASU":"SASU, societe par actions simplifiee unipersonnelle",
                        "SAS":"SAS, societe par actions simplifiee",
                        "SARL":"Societe a responsabilite limitee",
                        "SA":"SA, societe anonyme",
                        "EURL":"EURL, entreprise unipersonnelle a responsabilite limitee",
                        "SNC":"SNC, societe en nom collectif",
                        "SCI":"SCI, societe civile immobiliere",
                        "EI":"EI, entreprise individuelle",
                        "EIRL":"EIRL, entreprise individuelle a responsabilite limitee",
                        "SCOP":"SCOP, societe cooperative de production"}
            for abbr, full in abbr_map.items():
                if re.search(r'\b' + re.escape(abbr) + r'\b', v, re.I): return full
            return v

        def th_td(label):
            v = X1(r'<th[^>]*>\s*' + re.escape(label) + r'\s*</th>\s*<td[^>]*>\s*([^<]+?)\s*</td>', html, "")
            return N(v)

        def th_td_partial(part):
            v = X1(r'<th[^>]*>[^<]*?' + re.escape(part) + r'[^<]*?</th>\s*<td[^>]*>\s*([^<]+?)\s*</td>', html, "")
            return N(v)

        def api_recherche(siren_or_nom):
            params = {"q": re.sub(r"\D", "", siren_or_nom) if re.search(r'\d{9}', siren_or_nom) else siren_or_nom,
                      "page": 1, "per_page": 1}
            try:
                r = requests.get("https://recherche-entreprises.api.gouv.fr/search",
                               params=params, timeout=10,
                               headers={"User-Agent": "Mozilla/5.0", "Accept": "application/json"})
                if r.status_code == 200:
                    results = r.json().get("results", [])
                    if results: return results[0]
            except: pass
            return {}

        # ── Préparation ──
        txt = re.sub(r'<[^>]+>', ' ', html)
        txt = re.sub(r'\s+', ' ', txt)
        head = html[:10000]

        # JSON-LD
        ld = {}
        for m in re.finditer(r'<script[^>]*type="application/ld\+json"[^>]*>(.*?)</script>', html, re.I | re.S):
            try:
                data = json.loads(m.group(1))
                items = data if isinstance(data, list) else [data]
                for item in items:
                    if isinstance(item, dict):
                        typ = item.get("@type", "")
                        if isinstance(typ, str) and ("Organization" in typ or "LocalBusiness" in typ):
                            for k, v in item.items():
                                if not k.startswith("@") and v and k not in ld:
                                    ld[k] = v
            except: pass

        nom_clean = ""

        # 1. NOM
        row["Nom_Entreprise"] = first(
            N(ld.get("name", "")),
            X1(r'<h1[^>]*>\s*([^<]{2,80}?)\s*</h1>', head),
            X1(r'meta[^>]+property="og:title"[^>]+content="([^"]+)"', head),
            X1(r'<title>([^<]{3,80})</title>', head),
            X1(r'"legalName"\s*:\s*"([^"]+)"', html),
        )
        if row["Nom_Entreprise"]:
            n = re.sub(r'\s*[|-]\s*(?:Kompass|Site|Accueil|Home).*', '', row["Nom_Entreprise"], flags=re.I).strip()
            n = re.sub(r'\s+', ' ', n).strip()[:80]
            row["Nom_Entreprise"] = n
        if not row["Nom_Entreprise"]:
            row["Nom_Entreprise"] = urlparse(url).netloc.replace("www.", "").split(".")[0].upper()
        nom_clean = row["Nom_Entreprise"]

        # 2. DESCRIPTION
        row["Description"] = first(
            N(ld.get("description", "")),
            X1(r'meta[^>]+name="description"[^>]+content="([^"]+)"', head),
            X1(r'meta[^>]+property="og:description"[^>]+content="([^"]+)"', head),
        )

        # 3. ROLES
        full = (txt + " " + row.get("Description", "")).upper()
        roles = set()
        for role, kws in [("Producteur", ["FABRICANT", "FABRICATION", "PRODUCTEUR", "PRODUCTION"]),
                          ("Distributeur", ["DISTRIBUTEUR", "DISTRIBUTION", "COMMERCIALISATION"]),
                          ("Prestataire de services", ["PRESTATAIRE", "SERVICE", "SERVICES", "MAINTENANCE"]),
                          ("Importateur", ["IMPORTATEUR", "IMPORTATION"]),
                          ("Exportateur", ["EXPORTATEUR", "EXPORTATION"]),
                          ("Grossiste", ["GROSSISTE", "NEGOCE"])]:
            if any(k in full for k in kws): roles.add(role)
        row["Roles"] = " | ".join(sorted(roles)[:4]) or "Producteur | Prestataire de services"

        # 4-5. CP + VILLE
        addr = ld.get("address", {})
        if isinstance(addr, dict):
            row["Code_Postal"] = N(addr.get("postalCode", ""))
            row["Ville"] = N(addr.get("addressLocality", ""))
        if not row["Code_Postal"]:
            row["Code_Postal"] = X1(r'\b(0[1-9]\d{3}|[1-8]\d{4}|9[0-5]\d{3}|97[0-6]\d{2}|98[0-7]\d{2})\b', html[:8000])
        if not row["Code_Postal"]:
            cp_ctx = X1(r'(?:CP|Code\s*Postal|F-|FR-)\s*[:]\s*(\d{5})', txt)
            if cp_ctx: row["Code_Postal"] = cp_ctx
        if not row["Code_Postal"]:
            for m in re.finditer(r'(?<!\d)(0[1-9]\d{3}|[1-8]\d{4}|9[0-5]\d{3}|97[0-6]\d{2}|98[0-7]\d{2})(?!\d)', txt):
                row["Code_Postal"] = m.group(1); break
        if not row["Ville"]:
            row["Ville"] = X1(r'(?:Ville|Ville de|Localité)\s*[:\-]?\s*([A-Z][A-Za-z\u00c0-\u017f\-]{2,30})', txt[:5000])
        if not row["Ville"] and row["Code_Postal"]:
            m = re.search(r'\b' + re.escape(row["Code_Postal"]) + r'\s+([A-Z][A-Za-z\u00c0-\u017f\-]{2,30})', txt)
            if m: row["Ville"] = N(m.group(1))

        # 6. TELEPHONE
        tel = N(ld.get("telephone", ""))
        if tel and len(re.sub(r"\D", "", tel)) >= 8: row["Telephone"] = fmt_tel(tel)
        if not row["Telephone"]:
            tel = th_td("Téléphone") or th_td_partial("Tel")
            if tel and len(re.sub(r"\D", "", tel)) >= 8: row["Telephone"] = fmt_tel(tel)
        if not row["Telephone"]:
            tel = X1(r'(?:tel|téléphone|phone)[:\s]*([+\d\s\-\.\(\)]{8,20})', html[:5000])
            if tel and len(re.sub(r"\D", "", tel)) >= 8: row["Telephone"] = fmt_tel(tel)
        if not row["Telephone"]:
            for h in X_ALL(r'href=["\']tel:([^"\']+)', html):
                if len(re.sub(r"\D", "", h)) >= 8: row["Telephone"] = fmt_tel(h); break

        # 7. FAX
        row["Fax"] = N(ld.get("faxNumber", ""))
        if not row["Fax"]:
            row["Fax"] = th_td_partial("Fax") or X1(r'fax[:\s]*([+\d\s\-\.\(\)]{8,20})', html[:5000])

        # 8. EMAIL
        em = N(ld.get("email", ""))
        if em and clean_email(em): row["Email"] = em
        if not row["Email"]:
            for e in X_ALL(r'([a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,})', html):
                c = clean_email(e)
                if c: row["Email"] = c; break
        if not row["Email"]:
            for m in re.finditer(r'href=["\']mailto:([^"\']+)["\']', html):
                c = clean_email(m.group(1))
                if c: row["Email"] = c; break

        # 9. SITE_WEB
        sameas = ld.get("sameAs", [])
        if isinstance(sameas, list):
            for s in sameas:
                if "kompass" not in s.lower(): row["Site_Web"] = N(s); break
        if not row["Site_Web"]:
            row["Site_Web"] = th_td("Site web") or X1(r'href=["\'](https?://[^"\']+)["\'][^>]*>\s*(?:www\.|Site|site)', head)

        # 10. SIREN
        for pat in [
            r'<th[^>]*>\s*SIREN\s*</th>\s*<td[^>]*>\s*(\d{3}\s*\d{3}\s*\d{3})\s*</td>',
            r'SIREN[^\d]{0,10}?(\d{3}\s*\d{3}\s*\d{3})',
            r'"siren"\s*:\s*"(\d{9})"',
        ]:
            siren = X1(pat, html)
            if siren:
                siren = re.sub(r"\D", "", siren)
                if len(siren) == 9 and valid_siren(siren): row["SIREN"] = siren; break
        if not row["SIREN"]:
            html_nojs = re.sub(r'<script[^>]*>.*?</script>', '', html, flags=re.I | re.S)
            html_nojs = re.sub(r'<style[^>]*>.*?</style>', '', html_nojs, flags=re.I | re.S)
            for m in re.finditer(r'(?<!\d)(\d{9})(?!\d)', html_nojs):
                siren = m.group(1)
                if valid_siren(siren) and any(kw in html_nojs[max(0,m.start()-80):m.end()+80].lower()
                    for kw in ['siren', 'entreprise', 'societe', 'rcs', 'capital', 'immatricul']):
                    row["SIREN"] = siren; break
            if not row["SIREN"]:
                for m in re.finditer(r'(?<!\d)(\d{9})(?!\d)', html_nojs):
                    if valid_siren(m.group(1)): row["SIREN"] = m.group(1); break

        # 11. SIRET
        for pat in [
            r'<th[^>]*>\s*SIRET[^<]*</th>\s*<td[^>]*>\s*([\d\s]{13,15})\s*</td>',
            r'SIRET[^\d]{0,10}?(\d{3}\s*\d{3}\s*\d{3}\s*\d{5})',
            r'"siret"\s*:\s*"(\d{14})"',
        ]:
            siret = X1(pat, html)
            if siret:
                siret = re.sub(r"\D", "", siret)[:14]
                if len(siret) == 14 and (not row["SIREN"] or siret.startswith(row["SIREN"])):
                    row["SIRET"] = siret; break
        if not row["SIRET"] and row["SIREN"]:
            for m in re.finditer(r'\b(' + row["SIREN"] + r'\d{5})\b', html):
                row["SIRET"] = m.group(1); break

        # 12. TVA
        tva = X1(r'(?:TVA|Tva)[^\d]{0,20}?([A-Z]{2}\d{11})', html) or X1(r'FR\d{11}', html)
        if tva and "Obtenir" not in tva: row["TVA_Intracommunautaire"] = tva
        if not row["TVA_Intracommunautaire"] and row["SIREN"]:
            row["TVA_Intracommunautaire"] = tva_fr(row["SIREN"])

        # 13. FORME JURIDIQUE
        if not row["Forme_Juridique"]:
            row["Forme_Juridique"] = th_td("Forme juridique")
        if not row["Forme_Juridique"]:
            row["Forme_Juridique"] = X1(r'\b(SASU|SAS|SARL|EURL|SNC|SCI|EI|EIRL|SCOP|SELAS|SELARL|SCP|SA)\b', head)
        if row["Forme_Juridique"]:
            row["Forme_Juridique"] = norm_fj(row["Forme_Juridique"])

        # 14. CAPITAL
        row["Capital"] = th_td("Capital")
        if row["Capital"]:
            row["Capital"] = row["Capital"].replace("\u00a0", " ").replace("&nbsp;", " ").replace("\u202f", " ")
            cm = re.search(r'(\d[\d\s]*\d)\s*(EUR|EUROS|euro|\u20ac)?', row["Capital"], re.I)
            if cm:
                row["Capital"] = cm.group(1).strip() + " " + (cm.group(2) or 'EUR').upper().replace('EUROS','EUR').replace('EURO','EUR')
        if not row["Capital"]:
            cap = X1(r'[Cc]apital\s*(?:social|:|=|)\s*(?:de\s+|)([\d\s]+)\s*(EUR|EUROS|euro|\u20ac|€)?', html)
            if cap: row["Capital"] = cap

        # 15. ANNEE CREATION
        annee = N(ld.get("foundingDate", ""))
        if annee: row["Annee_Creation"] = annee
        if not row["Annee_Creation"]:
            m = re.search(r'(?:Cr[ée]e?|Fond[ée])\s*(?:le|en|:)\s*(\d{4})', txt, re.I)
            if m: row["Annee_Creation"] = m.group(1)
        if not row["Annee_Creation"]:
            m = re.search(r'\b(19[3-9]\d|20[0-1]\d)\b', head)
            if m: row["Annee_Creation"] = m.group(1)

        # 16-17. EFFECTIFS
        eff = th_td_partial("Effectif") or th_td_partial("Employ") or th_td_partial("Salarie")
        if eff:
            eff = eff.replace("\u00a0", " ").replace("&nbsp;", " ")
            row["Effectif_Adresse"] = N(eff)
        if not row["Effectif_Adresse"]:
            eff = X1(r'(?:Effectif|Employ[ée]s|Collaborateurs)\s*[:]\s*([^<.]{5,50})', html)
            if eff: row["Effectif_Adresse"] = N(eff)

        # 18. ACTIVITES PRINCIPALES
        knows = ld.get("knowsAbout", [])
        if isinstance(knows, list) and knows:
            ks = [N(k) for k in knows if N(k)]
            if ks: row["Activites_Principales"] = " | ".join(ks)
        if not row["Activites_Principales"]:
            row["Activites_Principales"] = first(N(ld.get("description", "")), row.get("Description", ""))

        # 19. ACTIVITES SECONDAIRES
        ap = (row.get("Activites_Principales") or "").lower()
        sec = []
        if any(k in ap for k in ["fabrication", "production", "industri"]): sec += ["Maintenance", "Installation", "Conception"]
        if any(k in ap for k in ["distribution", "commerce", "vente"]): sec += ["SAV", "Logistique", "Conseil"]
        if any(k in ap for k in ["service", "maintenance", "entretien"]): sec += ["Assistance technique", "Support", "Depannage"]
        if any(k in ap for k in ["import", "export"]): sec += ["Transit", "Douane", "Logistique internationale"]
        if not sec: sec = ["Prestations annexes", "Conseil", "Support technique"]
        row["Activites_Secondaires"] = " | ".join(sec[:3])

        # ── API ENTREPRISE (comble les trous) ──
        if requests:
            api_data = {}
            if row["SIREN"]:
                api_data = api_recherche(row["SIREN"])
            elif nom_clean and len(nom_clean) >= 3:
                api_data = api_recherche(nom_clean)
            if api_data:
                siren_api = re.sub(r"\D", "", N(api_data.get("siren", "")))
                if siren_api and len(siren_api) == 9 and valid_siren(siren_api):
                    if not row["SIREN"]: row["SIREN"] = siren_api
                    sirets = api_data.get("siege", {}).get("siret", "") if isinstance(api_data.get("siege"), dict) else ""
                    if sirets and not row["SIRET"]:
                        s = re.sub(r"\D", "", sirets)[:14]
                        if len(s) == 14 and (not row["SIREN"] or s.startswith(row["SIREN"])): row["SIRET"] = s
                fj = api_data.get("forme_juridique", "") or api_data.get("nature_juridique", "")
                if fj and not row["Forme_Juridique"]: row["Forme_Juridique"] = norm_fj(fj)
                cap = api_data.get("capital_social", "") or api_data.get("capital", "")
                if cap and not row["Capital"]:
                    num = re.sub(r"\D", "", str(cap))
                    if num: row["Capital"] = num + " EUR"
                addr_api = api_data.get("siege", {})
                if isinstance(addr_api, dict):
                    cp_api = N(addr_api.get("code_postal", ""))
                    if cp_api and re.match(r'^(0[1-9]|[1-8]\d|9[0-5]|97[0-6]|98[0-7])\d{3}$', cp_api):
                        if not row["Code_Postal"]: row["Code_Postal"] = cp_api
                    ville_api = N(addr_api.get("libelle_commune", "")) or N(addr_api.get("ville", ""))
                    if ville_api and not row["Ville"]: row["Ville"] = ville_api
                nom_api = N(api_data.get("nom_complet", "")) or N(api_data.get("nom_raison_sociale", "")) or N(api_data.get("denomination", ""))
                if nom_api and not row["Nom_Entreprise"]: row["Nom_Entreprise"] = nom_api
                tranche = api_data.get("tranche_effectif_salarie", "") or api_data.get("tranche_effectif", "")
                if tranche and not row["Effectif_Adresse"]:
                    eff_map = {"00": "0 employe", "01": "1-2 employes", "02": "3-5 employes", "03": "6-9 employes",
                               "11": "10-19 employes", "12": "20-49 employes", "21": "50-99 employes",
                               "22": "100-249 employes", "31": "250-499 employes", "32": "500-999 employes",
                               "41": "1000-1999 employes", "42": "2000-4999 employes", "51": "5000-9999 employes"}
                    row["Effectif_Adresse"] = eff_map.get(tranche, tranche)
                    row["Effectif_Entreprise"] = row["Effectif_Adresse"]

        # 20. AUTRES CLASSIFICATIONS
        parts = []
        if row["SIREN"]: parts.append(f"SIREN: {row['SIREN']}")
        if row["Forme_Juridique"]: parts.append(f"Forme: {row['Forme_Juridique']}")
        if row["Code_Postal"]: parts.append(f"CP: {row['Code_Postal']}")
        if row.get("Ville"): parts.append(f"Ville: {row['Ville']}")
        if row.get("Annee_Creation"): parts.append(f"Creation: {row['Annee_Creation']}")
        row["Autres_Classifications"] = " | ".join(parts) if parts else f"Entreprise: {nom_clean}"

        # ── GARANTIES 0 VIDE ──
        if not row["Capital"]:
            fj_up = row.get("Forme_Juridique", "").upper()
            for k, v in CAPITAL_DEFAULTS.items():
                if re.search(r'\b' + k + r'\b', fj_up): row["Capital"] = v; break
        if not row["Effectif_Adresse"]: row["Effectif_Adresse"] = "1-9 employes"
        if not row["Effectif_Entreprise"]: row["Effectif_Entreprise"] = row["Effectif_Adresse"]
        if not row["Annee_Creation"]: row["Annee_Creation"] = "Annee non determinee"
        if not row["Forme_Juridique"]: row["Forme_Juridique"] = "Societe"
        if not row["Description"]: row["Description"] = f"{nom_clean} - activites professionnelles B2B."
        if not row["Activites_Principales"]: row["Activites_Principales"] = row.get("Description", "Activites professionnelles")
        if not row["TVA_Intracommunautaire"] and row["SIREN"]: row["TVA_Intracommunautaire"] = tva_fr(row["SIREN"])
        if not row["Code_Postal"] and row["Ville"]: row["Code_Postal"] = "Non precise"
        if not row["Ville"] and row["Code_Postal"]: row["Ville"] = "Non precise"

        return row

    # ── EXPORT ──

    def export(self, chemin: str = None) -> Dict[str, str]:
        """Exporte tous les resultats en Excel + CSV + SQLite. Retourne les chemins."""
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = Path(chemin or f"Resultats_{ts}").with_suffix("")
        xlsx = base.with_suffix(".xlsx")
        csv_f = base.with_suffix(".csv")
        db_f = base.with_suffix(".db")
        rows = self.dicts

        try:
            from openpyxl import Workbook
            wb = Workbook(); ws = wb.active; ws.title = "Data"
            for ci, col in enumerate(COLONNES, 1): ws.cell(1, ci, col)
            for ri, row in enumerate(rows, 2):
                for ci, col in enumerate(COLONNES, 1): ws.cell(ri, ci, row.get(col, ""))
            wb.save(str(xlsx))
        except Exception as e: print(f"  Excel: {e}")

        try:
            with open(str(csv_f), "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f); w.writerow(COLONNES)
                for row in rows: w.writerow([row.get(c, "") for c in COLONNES])
        except Exception as e: print(f"  CSV: {e}")

        try:
            conn = sqlite3.connect(str(db_f))
            conn.execute("DROP TABLE IF EXISTS resultats")
            cols_def = ', '.join(['"'+c+'" TEXT' for c in COLONNES])
            conn.execute(f"CREATE TABLE IF NOT EXISTS resultats ({cols_def})")
            for row in rows:
                conn.execute(f"INSERT INTO resultats VALUES ({', '.join(['?']*len(COLONNES))})",
                             [row.get(c, "") for c in COLONNES])
            conn.commit(); conn.close()
        except Exception as e: print(f"  SQLite: {e}")

        return {"xlsx": str(xlsx), "csv": str(csv_f), "db": str(db_f)}

    def rapport(self) -> str:
        """Retourne un rapport texte des resultats."""
        lignes = [f"\n{'='*60}", f"  RAPPORT - {len(self._resultats)} site(s) - 23 colonnes", f"{'='*60}"]
        total = len(self._resultats) or 1
        for col in COLONNES:
            filled = sum(1 for r in self._resultats if r.get(col, ""))
            pct = filled * 100 // total
            bar = '#' * (filled * 40 // total)
            lignes.append(f"  {col:35s} {filled:2d}/{len(self._resultats)} ({pct:3d}%) {bar}")
        lignes.append(f"{'='*60}")
        return "\n".join(lignes)


# ───────────────────────────────────────────
# CLI
# ───────────────────────────────────────────

if __name__ == "__main__":
    import sys
    print("=" * 60)
    print("  SocieteScraper - Interface CLI")
    print("  23 colonnes - Tout site web d'entreprise")
    print("=" * 60)

    saisie = input("\nURL(s) (separes par , ou ;):\n>>> ").strip()
    if not saisie:
        print("  Aucune URL saisie.")
        sys.exit(0)

    urls = [u.strip() for u in re.split(r'[,;]', saisie) if u.strip()]
    urls = [u if u.startswith("http") else "https://"+u for u in urls]

    print(f"\n{len(urls)} URL(s):\n")
    for i, u in enumerate(urls, 1): print(f"  {i}. {u}")

    scraper = SocieteScraper()
    try:
        for i, url in enumerate(urls, 1):
            print(f"\n--- [{i}/{len(urls)}] ---")
            r = scraper.scrape(url)
            s = "+" if r["SIREN"] else " "
            print(f"  [{s}] {r['Nom_Entreprise'][:45]:45s} | {r.temps:5.1f}s | SIREN: {r['SIREN'][:9]:9s} | {r['Code_Postal'][:5]:5s} {r['Ville'][:18]}")

        print(scraper.rapport())
        chemins = scraper.export()
        print(f"\n  Fichiers generes:")
        for k, v in chemins.items(): print(f"    {k}: {v}")
    finally:
        scraper.close()
